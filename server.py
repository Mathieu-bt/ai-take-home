import base64
import io
import json
import os
import re
from copy import copy
from datetime import datetime
from typing import Any, Dict, List

import openpyxl
import requests
from dotenv import load_dotenv
from fastapi import FastAPI, File, HTTPException, UploadFile
from fastapi.responses import JSONResponse
from fastapi.staticfiles import StaticFiles

from db import find_best_client, get_gl_mapping, get_output_default, init_db, list_seed_data


load_dotenv()
init_db()

GEMINI_API_KEY = os.getenv("GEMINI_API_KEY", "")
GEMINI_MODEL = os.getenv("GEMINI_MODEL", "gemini-2.0-flash")
LAYOUT_TEMPLATE_PATH = os.getenv("LAYOUT_TEMPLATE_PATH", "./Heron Data - AR ingestion layout.xlsx")

OUTPUT_COLUMNS = [
  "upload",
  "voucher_no",
  "GL lookup",
  "trans_type",
  "GL account name",
  "Client_id_lookup",
  "Client_name",
  "period",
  "Dimension_id_lookup",
  "Dimension_name",
  "Suppl/cust_name",
  "inv_ref",
  "invoice_date",
  "currency",
  "cur_amount",
  "description",
  "notes",
]

INVOICE_JSON_SCHEMA: Dict[str, Any] = {
  "type": "OBJECT",
  "required": [
    "invoice_number",
    "invoice_date",
    "due_date",
    "currency",
    "customer_name",
    "service_type",
    "service_description",
    "service_period_label",
    "net_amount",
    "tax_label",
    "tax_amount",
    "total_amount",
    "is_tax_invoice",
    "confidence",
  ],
  "properties": {
    "invoice_number": {"type": "STRING"},
    "invoice_date": {"type": "STRING"},
    "due_date": {"type": "STRING"},
    "currency": {"type": "STRING"},
    "customer_name": {"type": "STRING"},
    "service_type": {"type": "STRING"},
    "service_description": {"type": "STRING"},
    "service_period_label": {"type": "STRING"},
    "net_amount": {"type": "NUMBER"},
    "tax_label": {"type": "STRING"},
    "tax_amount": {"type": "NUMBER"},
    "total_amount": {"type": "NUMBER"},
    "is_tax_invoice": {"type": "BOOLEAN"},
    "confidence": {"type": "NUMBER"},
  },
}

PROMPT = """
You are extracting structured accounting data from a customer invoice PDF.
OCR the full document and return ONLY valid JSON that matches the schema.

Rules:
- Date format must be DD/MM/YYYY.
- Amount fields must be numeric (not strings).
- customer_name = the billed customer receiving the invoice.
- service_period_label should look like "January 2026", "February 2026", etc.
- If no tax is present, set tax_amount to 0 and tax_label to "No tax".
- confidence is between 0 and 1.
"""

app = FastAPI(title="AR Ingestion Demo API")


def _parse_ddmmyyyy_to_yyyymmdd(d: str) -> str:
  match = re.match(r"^(\d{2})/(\d{2})/(\d{4})$", str(d or "").strip())
  if not match:
    return ""
  return f"{match.group(3)}{match.group(2)}{match.group(1)}"


def _period_from_yyyymmdd(yyyymmdd: str) -> str:
  if not yyyymmdd or len(yyyymmdd) < 6:
    return ""
  return yyyymmdd[:6]


def _period_from_service_label(service_period_label: str) -> str:
  value = str(service_period_label or "").strip()
  if not value:
    return ""
  value_low = value.lower()
  month_map = {
    "january": "01",
    "february": "02",
    "march": "03",
    "april": "04",
    "may": "05",
    "june": "06",
    "july": "07",
    "august": "08",
    "september": "09",
    "october": "10",
    "november": "11",
    "december": "12",
  }
  year_match = re.search(r"(20\d{2})", value_low)
  if not year_match:
    return ""
  year = year_match.group(1)
  month = ""
  for name, mm in month_map.items():
    if name in value_low:
      month = mm
      break
  return f"{year}{month}" if month else ""


def _build_description(client_name: str, dimension_name: str, period: str) -> str:
  return " ".join(part for part in [client_name, dimension_name, period] if part).strip()


def _safe_json_parse(text: str) -> Dict[str, Any]:
  cleaned = text.strip()
  cleaned = re.sub(r"^```json", "", cleaned).strip()
  cleaned = re.sub(r"^```", "", cleaned).strip()
  cleaned = re.sub(r"```$", "", cleaned).strip()
  return json.loads(cleaned)


def extract_invoice_with_gemini(pdf_bytes: bytes, filename: str) -> Dict[str, Any]:
  if not GEMINI_API_KEY:
    raise HTTPException(status_code=500, detail="Missing GEMINI_API_KEY in .env")

  url = (
    f"https://generativelanguage.googleapis.com/v1beta/models/{GEMINI_MODEL}:generateContent"
  )
  payload = {
    "contents": [
      {
        "parts": [
          {"text": PROMPT},
          {
            "inline_data": {
              "mime_type": "application/pdf",
              "data": base64.b64encode(pdf_bytes).decode("utf-8"),
            }
          },
        ]
      }
    ],
    "generationConfig": {
      "temperature": 0.1,
      "responseMimeType": "application/json",
      "responseSchema": INVOICE_JSON_SCHEMA,
    },
  }

  response = requests.post(
    url,
    json=payload,
    headers={"x-goog-api-key": GEMINI_API_KEY},
    timeout=90,
  )
  if response.status_code >= 400:
    raise HTTPException(
      status_code=500,
      detail=f"Gemini API error for {filename}: {response.status_code}",
    )

  data = response.json()
  try:
    text = data["candidates"][0]["content"]["parts"][0]["text"]
    parsed = _safe_json_parse(text)
  except Exception as exc:
    raise HTTPException(status_code=500, detail=f"Invalid Gemini response for {filename}: {exc}") from exc

  parsed["source_file"] = filename
  return parsed


def build_rows_from_invoice(extracted: Dict[str, Any], voucher_no: int) -> List[Dict[str, Any]]:
  client_match = find_best_client(extracted.get("customer_name", ""))
  notes: List[str] = []
  upload_flag = "yes"
  if client_match is None:
    upload_flag = "no"
    notes.append("CLIENT NOT IN MAPPING")
    client_id_lookup = "UNKNOWN"
    client_name = "UNKNOWN"
    supplier_customer_name = "UNKNOWN"
    dimension_id = "UNKNOWN"
    dimension_name = "UNKNOWN"
  else:
    client_id_lookup = client_match["client_code"]
    client_name = client_match.get("canonical_name", "UNKNOWN")
    supplier_customer_name = client_match.get("supplier_customer_name", "UNKNOWN")
    dimension_id = client_match.get("default_dimension_id") or "UNKNOWN"
    dimension_name = client_match.get("default_dimension_name") or "UNKNOWN"

  invoice_date_fmt = _parse_ddmmyyyy_to_yyyymmdd(extracted.get("invoice_date", ""))
  period = _period_from_service_label(extracted.get("service_period_label", "")) or _period_from_yyyymmdd(invoice_date_fmt)
  if not period:
    notes.append("Missing service period label; period fallback failed.")

  dim_id_header = "UNKNOWN" if upload_flag == "no" else ""
  dim_name_header = "UNKNOWN" if upload_flag == "no" else ""

  ar = get_gl_mapping("AR")
  service = get_gl_mapping("SERVICE")
  tax = get_gl_mapping("TAX")

  total_amount = float(extracted.get("total_amount", 0) or 0)
  net_amount = float(extracted.get("net_amount", 0) or 0)
  tax_amount = float(extracted.get("tax_amount", 0) or 0)
  currency = (extracted.get("currency") or "").upper()
  invoice_number = extracted.get("invoice_number", "")

  rows = [
    {
      "upload": upload_flag,
      "voucher_no": voucher_no,
      "GL lookup": ar["gl_lookup"],
      "trans_type": ar["trans_type"],
      "GL account name": ar["gl_account_name"],
      "Client_id_lookup": client_id_lookup,
      "Client_name": client_name,
      "period": period,
      "Dimension_id_lookup": dim_id_header,
      "Dimension_name": dim_name_header,
      "Suppl/cust_name": supplier_customer_name,
      "inv_ref": invoice_number,
      "invoice_date": invoice_date_fmt,
      "currency": currency,
      "cur_amount": round(total_amount, 2),
      "description": _build_description(client_name, "AR", period),
      "notes": "; ".join(notes),
    },
    {
      "upload": upload_flag,
      "voucher_no": voucher_no,
      "GL lookup": service["gl_lookup"],
      "trans_type": service["trans_type"],
      "GL account name": service["gl_account_name"],
      "Client_id_lookup": client_id_lookup,
      "Client_name": client_name,
      "period": period,
      "Dimension_id_lookup": dimension_id,
      "Dimension_name": dimension_name,
      "Suppl/cust_name": supplier_customer_name,
      "inv_ref": invoice_number,
      "invoice_date": invoice_date_fmt,
      "currency": currency,
      "cur_amount": round(-net_amount, 2),
      "description": _build_description(client_name, dimension_name or "Service", period),
      "notes": "; ".join(notes),
    },
  ]

  tax_threshold = float(get_output_default("tax_row_threshold", "0.0001"))
  if abs(tax_amount) > tax_threshold:
    rows.append(
      {
        "upload": upload_flag,
        "voucher_no": voucher_no,
        "GL lookup": tax["gl_lookup"],
        "trans_type": tax["trans_type"],
        "GL account name": tax["gl_account_name"],
        "Client_id_lookup": client_id_lookup,
        "Client_name": client_name,
        "period": period,
        "Dimension_id_lookup": dim_id_header,
        "Dimension_name": dim_name_header,
        "Suppl/cust_name": supplier_customer_name,
        "inv_ref": invoice_number,
        "invoice_date": invoice_date_fmt,
        "currency": currency,
        "cur_amount": round(-tax_amount, 2),
        "description": _build_description(client_name, "Tax", period),
        "notes": "; ".join(notes),
      }
    )

  return rows


def fill_layout_template(rows: List[Dict[str, Any]]) -> bytes:
  template_path = os.path.abspath(LAYOUT_TEMPLATE_PATH)
  if not os.path.exists(template_path):
    raise HTTPException(status_code=500, detail=f"Layout template not found: {template_path}")

  wb = openpyxl.load_workbook(template_path)
  ws = wb["Layout"] if "Layout" in wb.sheetnames else wb.active

  start_row = 15
  start_col = 1
  end_col = len(OUTPUT_COLUMNS)

  # clear existing values in data block while keeping formatting
  for r in range(start_row, ws.max_row + 1):
    for c in range(start_col, end_col + 1):
      ws.cell(row=r, column=c).value = None

  template_style_row = start_row
  for idx, row in enumerate(rows):
    target_row = start_row + idx
    if target_row != template_style_row:
      for c in range(start_col, end_col + 1):
        src = ws.cell(row=template_style_row, column=c)
        tgt = ws.cell(row=target_row, column=c)
        tgt._style = copy(src._style)
        tgt.number_format = src.number_format
        tgt.font = copy(src.font)
        tgt.fill = copy(src.fill)
        tgt.border = copy(src.border)
        tgt.alignment = copy(src.alignment)
        tgt.protection = copy(src.protection)
      if ws.row_dimensions.get(template_style_row):
        ws.row_dimensions[target_row].height = ws.row_dimensions[template_style_row].height

    for col_idx, field in enumerate(OUTPUT_COLUMNS, start=start_col):
      ws.cell(row=target_row, column=col_idx).value = row.get(field, "")

  bio = io.BytesIO()
  wb.save(bio)
  bio.seek(0)
  return bio.read()


@app.get("/api/health")
def api_health() -> Dict[str, Any]:
  return {"ok": True, "gemini_model": GEMINI_MODEL, "has_api_key": bool(GEMINI_API_KEY)}


@app.get("/api/mappings")
def api_mappings() -> Dict[str, Any]:
  return list_seed_data()


@app.post("/api/process")
async def api_process(files: List[UploadFile] = File(...)) -> JSONResponse:
  if not files:
    raise HTTPException(status_code=400, detail="No files uploaded.")

  parsed_invoices = []
  all_rows = []

  for idx, file in enumerate(files, start=0):
    content = await file.read()
    extracted = extract_invoice_with_gemini(content, file.filename)
    parsed_invoices.append(extracted)
    all_rows.extend(build_rows_from_invoice(extracted, voucher_no=idx))

  workbook_bytes = fill_layout_template(all_rows)
  workbook_b64 = base64.b64encode(workbook_bytes).decode("utf-8")

  return JSONResponse(
    content={
      "parsed_invoices": parsed_invoices,
      "journal_rows": all_rows,
      "output_columns": OUTPUT_COLUMNS,
      "workbook_base64": workbook_b64,
      "workbook_filename": f"Heron_AR_ingestion_output_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
    }
  )


app.mount("/", StaticFiles(directory=".", html=True), name="static")
